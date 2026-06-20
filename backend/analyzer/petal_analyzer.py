import cv2
import numpy as np
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


@dataclass
class AnalysisResult:
    vein_count: int
    serration_area: float
    serration_count: int
    petal_area: float
    petal_perimeter: float
    circularity: float
    shape_type: str = ""
    shape_score: dict = field(default_factory=dict)
    contours: List[List[Tuple[int, int]]] = field(default_factory=list)
    vein_points: List[Tuple[int, int]] = field(default_factory=list)
    serration_points: List[Tuple[int, int]] = field(default_factory=list)
    original_width: int = 0
    original_height: int = 0

    def to_dict(self):
        return {
            "vein_count": int(self.vein_count),
            "serration_area": float(self.serration_area),
            "serration_count": int(self.serration_count),
            "petal_area": float(self.petal_area),
            "petal_perimeter": float(self.petal_perimeter),
            "circularity": float(self.circularity),
            "shape_type": self.shape_type,
            "shape_score": {k: float(v) for k, v in self.shape_score.items()},
            "contours": [[(int(point[0]), int(point[1])) for point in contour] for contour in self.contours],
            "vein_points": [(int(point[0]), int(point[1])) for point in self.vein_points],
            "serration_points": [(int(point[0]), int(point[1])) for point in self.serration_points],
            "original_width": int(self.original_width),
            "original_height": int(self.original_height),
        }


class PetalAnalyzer:
    def __init__(self):
        self.min_petal_area_ratio = 0.005
        self.max_petal_area_ratio = 0.95
        
        self.min_vein_length_ratio = 0.025
        self.min_vein_solidity = 0.22
        self.vein_edge_margin_ratio = 0.04
        
        self.min_serration_depth_ratio = 0.008
        self.min_serration_width_ratio = 0.002

    def analyze(self, image_bytes: bytes) -> AnalysisResult:
        img_array = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

        if img is None:
            raise ValueError("无法解析图片")

        original_height, original_width = img.shape[:2]
        img_area = original_width * original_height

        preprocessed_gray = self._preprocess_image(img)
        binary = self._segment_petal(img, preprocessed_gray)
        cleaned = self._morphological_clean(binary)

        contours, hierarchy = self._find_contours(cleaned)
        petal_contour = self._find_petal_contour(contours, img_area)

        if petal_contour is None:
            return AnalysisResult(
                vein_count=0,
                serration_area=0,
                serration_count=0,
                petal_area=0,
                petal_perimeter=0,
                circularity=0,
                shape_type="未知",
                shape_score={"圆形": 0.0, "尖形": 0.0, "波浪形": 0.0},
                original_width=original_width,
                original_height=original_height,
            )

        petal_contour = self._smooth_contour(petal_contour, kernel_size=1)

        petal_area = cv2.contourArea(petal_contour)
        petal_perimeter = cv2.arcLength(petal_contour, True)

        circularity = 0
        if petal_perimeter > 0:
            circularity = 4 * np.pi * petal_area / (petal_perimeter * petal_perimeter)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        vein_count, vein_points = self._detect_veins(
            gray, petal_contour, img_area
        )
        serration_area, serration_count, serration_points = self._calculate_serration(
            petal_contour, petal_area
        )

        shape_type, shape_score = self._classify_shape(
            petal_contour, petal_area, petal_perimeter, circularity,
            serration_count, vein_count
        )

        result = AnalysisResult(
            vein_count=vein_count,
            serration_area=serration_area,
            serration_count=serration_count,
            petal_area=petal_area,
            petal_perimeter=petal_perimeter,
            circularity=circularity,
            shape_type=shape_type,
            shape_score=shape_score,
            contours=[petal_contour.reshape(-1, 2).tolist()],
            vein_points=vein_points,
            serration_points=serration_points,
            original_width=original_width,
            original_height=original_height,
        )

        return result

    def _preprocess_image(self, img: np.ndarray) -> np.ndarray:
        lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        l_eq = clahe.apply(l)
        
        lab_eq = cv2.merge([l_eq, a, b])
        bgr_eq = cv2.cvtColor(lab_eq, cv2.COLOR_LAB2BGR)
        
        gray = cv2.cvtColor(bgr_eq, cv2.COLOR_BGR2GRAY)
        
        denoised = cv2.bilateralFilter(gray, d=9, sigmaColor=75, sigmaSpace=75)
        
        return denoised

    def _correct_illumination(self, gray: np.ndarray) -> np.ndarray:
        kernel_size = max(gray.shape) // 50
        kernel_size = max(kernel_size, 15)
        if kernel_size % 2 == 0:
            kernel_size += 1
        
        struct_elem = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (kernel_size, kernel_size)
        )
        bg = cv2.morphologyEx(gray, cv2.MORPH_OPEN, struct_elem)
        bg = cv2.GaussianBlur(bg, (kernel_size * 2 + 1, kernel_size * 2 + 1), 0)
        
        bg = bg.astype(np.float32)
        gray_float = gray.astype(np.float32)
        
        ratio = gray_float / (bg + 1e-6)
        corrected = ratio * 128.0
        corrected = np.clip(corrected, 0, 255).astype(np.uint8)
        
        return corrected

    def _segment_petal(self, color_img: np.ndarray, gray: np.ndarray) -> np.ndarray:
        h, w = gray.shape
        
        hsv = cv2.cvtColor(color_img, cv2.COLOR_BGR2HSV)
        hsv_h, hsv_s, hsv_v = cv2.split(hsv)
        
        lab = cv2.cvtColor(color_img, cv2.COLOR_BGR2LAB)
        lab_l, lab_a, lab_b = cv2.split(lab)
        
        hue_mask = cv2.inRange(hsv, (120, 20, 50), (175, 255, 255))
        
        _, s_otsu = cv2.threshold(
            hsv_s, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        s_binary = cv2.bitwise_and(s_otsu, hue_mask)
        
        _, gray_otsu = cv2.threshold(
            gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU
        )
        
        combined = cv2.bitwise_and(s_binary, gray_otsu)
        
        a_normalized = cv2.normalize(lab_a, None, 0, 255, cv2.NORM_MINMAX)
        _, a_binary = cv2.threshold(
            a_normalized, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        a_binary = cv2.bitwise_and(a_binary, hue_mask)
        
        combined = cv2.bitwise_or(combined, a_binary)
        combined = cv2.bitwise_and(combined, gray_otsu)
        combined = cv2.bitwise_and(combined, hue_mask)
        
        edge_mask = np.ones((h, w), dtype=np.uint8) * 255
        border_size = int(min(h, w) * 0.02)
        edge_mask[:border_size, :] = 0
        edge_mask[-border_size:, :] = 0
        edge_mask[:, :border_size] = 0
        edge_mask[:, -border_size:] = 0
        
        combined = cv2.bitwise_and(combined, edge_mask)
        
        return combined

    def _morphological_clean(self, binary: np.ndarray) -> np.ndarray:
        kernel_small = np.ones((3, 3), np.uint8)
        kernel_medium = np.ones((5, 5), np.uint8)

        cleaned = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel_small)
        cleaned = cv2.morphologyEx(cleaned, cv2.MORPH_CLOSE, kernel_medium)

        contours, _ = cv2.findContours(
            cleaned, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )
        
        mask = np.zeros_like(cleaned)
        h, w = cleaned.shape[:2]
        img_area = h * w
        min_area = img_area * 0.001
        
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area > min_area:
                cv2.drawContours(mask, [cnt], -1, 255, -1)
        
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel_medium)
        mask = cv2.dilate(mask, kernel_small, iterations=1)

        return mask

    def _find_contours(self, binary: np.ndarray) -> Tuple[List[np.ndarray], np.ndarray]:
        contours, hierarchy = cv2.findContours(
            binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )
        return contours, hierarchy

    def _find_petal_contour(
        self, contours: List[np.ndarray], img_area: int
    ) -> np.ndarray:
        if not contours:
            return None

        min_area = img_area * self.min_petal_area_ratio
        max_area = img_area * self.max_petal_area_ratio

        candidates = []
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < min_area or area > max_area:
                continue

            x, y, w, h = cv2.boundingRect(cnt)
            aspect_ratio = float(w) / h if h > 0 else 0
            if aspect_ratio < 0.1 or aspect_ratio > 10:
                continue

            hull = cv2.convexHull(cnt)
            hull_area = cv2.contourArea(hull)
            solidity = area / hull_area if hull_area > 0 else 0
            if solidity < 0.2:
                continue

            perimeter = cv2.arcLength(cnt, True)
            circularity = 0
            if perimeter > 0:
                circularity = 4 * np.pi * area / (perimeter * perimeter)

            rect_area = w * h
            extent = area / rect_area if rect_area > 0 else 0

            score = area * (0.5 + circularity) * (0.5 + solidity) * (0.5 + extent)
            candidates.append((score, cnt))

        if not candidates:
            return max(contours, key=cv2.contourArea)

        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    def _smooth_contour(self, contour: np.ndarray, kernel_size: int = 3) -> np.ndarray:
        contour_float = contour.reshape(-1, 2).astype(np.float32)
        
        smoothed = np.copy(contour_float)
        n = len(contour_float)
        
        for i in range(n):
            total = np.zeros(2, dtype=np.float32)
            weight_sum = 0.0
            for j in range(-kernel_size, kernel_size + 1):
                idx = (i + j) % n
                weight = 1.0 - abs(j) / (kernel_size + 1)
                total += contour_float[idx] * weight
                weight_sum += weight
            smoothed[i] = total / weight_sum
        
        smoothed = smoothed.astype(np.int32).reshape(-1, 1, 2)
        return smoothed

    def _detect_veins(
        self, gray: np.ndarray, petal_contour: np.ndarray, img_area: int
    ) -> Tuple[int, List[Tuple[int, int]]]:
        h, w = gray.shape
        
        mask = np.zeros_like(gray)
        cv2.drawContours(mask, [petal_contour], -1, 255, -1)

        petal_area = cv2.contourArea(petal_contour)
        if petal_area < 500:
            return 0, []

        petal_size = np.sqrt(petal_area)
        min_vein_length = max(15, petal_size * self.min_vein_length_ratio)
        
        margin = max(5, int(petal_size * self.vein_edge_margin_ratio))
        inner_kernel = np.ones((margin * 2, margin * 2), np.uint8)
        inner_mask = cv2.erode(mask, inner_kernel, iterations=1)

        masked_gray = cv2.bitwise_and(gray, gray, mask=inner_mask)

        kernel_size = max(3, int(petal_size * 0.02))
        if kernel_size % 2 == 0:
            kernel_size += 1
        struct_elem = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (kernel_size, kernel_size)
        )
        
        blackhat = cv2.morphologyEx(masked_gray, cv2.MORPH_BLACKHAT, struct_elem)
        tophat = cv2.morphologyEx(masked_gray, cv2.MORPH_TOPHAT, struct_elem)
        vein_enhanced = cv2.add(blackhat, tophat)

        vein_enhanced = cv2.bitwise_and(vein_enhanced, vein_enhanced, mask=inner_mask)

        enhanced_pixels = vein_enhanced[inner_mask > 0]
        if len(enhanced_pixels) == 0:
            return 0, []
        
        mean_val = np.mean(enhanced_pixels)
        std_val = np.std(enhanced_pixels)
        threshold = mean_val + std_val * 0.6
        
        _, vein_binary = cv2.threshold(
            vein_enhanced, threshold, 255, cv2.THRESH_BINARY
        )
        vein_binary = cv2.bitwise_and(vein_binary, vein_binary, mask=inner_mask)

        close_kernel_size = max(3, int(petal_size * 0.01))
        if close_kernel_size % 2 == 0:
            close_kernel_size += 1
        close_kernel = cv2.getStructuringElement(
            cv2.MORPH_ELLIPSE, (close_kernel_size, close_kernel_size)
        )
        vein_binary = cv2.morphologyEx(vein_binary, cv2.MORPH_CLOSE, close_kernel)
        vein_binary = cv2.morphologyEx(vein_binary, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8))

        vein_contours, _ = cv2.findContours(
            vein_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
        )

        valid_veins = []
        for cnt in vein_contours:
            length = cv2.arcLength(cnt, False)
            if length < min_vein_length:
                continue

            area = cv2.contourArea(cnt)
            if area < max(8, petal_area * 0.00008):
                continue

            x, y, cw, ch = cv2.boundingRect(cnt)
            aspect = max(cw, ch) / min(cw, ch) if min(cw, ch) > 0 else 0
            if aspect < 1.05:
                continue

            hull = cv2.convexHull(cnt)
            hull_area = cv2.contourArea(hull)
            solidity = area / hull_area if hull_area > 0 else 0
            if solidity < self.min_vein_solidity:
                continue

            if area > 0:
                thinness = length * length / (4 * np.pi * area)
                if thinness < 1.5:
                    continue

            M = cv2.moments(cnt)
            if M["m00"] != 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])
                
                dist = cv2.pointPolygonTest(petal_contour, (cx, cy), True)
                if dist > margin * 0.2:
                    valid_veins.append((cnt, (cx, cy)))

        valid_veins.sort(key=lambda x: cv2.contourArea(x[0]), reverse=True)
        
        valid_veins = self._merge_duplicate_veins(valid_veins, petal_contour)
        
        max_veins = 100
        valid_veins = valid_veins[:max_veins]

        vein_points = [vp[1] for vp in valid_veins]

        return len(valid_veins), vein_points
    
    def _merge_duplicate_veins(
        self, veins: List, petal_contour: np.ndarray
    ) -> List:
        if len(veins) <= 1:
            return veins
        
        M = cv2.moments(petal_contour)
        if M["m00"] == 0:
            return veins
        center_x = int(M["m10"] / M["m00"])
        center_y = int(M["m01"] / M["m00"])
        
        petal_area = cv2.contourArea(petal_contour)
        petal_size = np.sqrt(petal_area)
        
        vein_angles = []
        vein_radii = []
        vein_areas = []
        for cnt, (cx, cy) in veins:
            angle = np.arctan2(cy - center_y, cx - center_x)
            if angle < 0:
                angle += 2 * np.pi
            vein_angles.append(angle)
            
            radius = np.sqrt((cx - center_x)**2 + (cy - center_y)**2)
            vein_radii.append(radius)
            vein_areas.append(cv2.contourArea(cnt))
        
        merged = []
        used = [False] * len(veins)
        
        angle_threshold_small = 0.15
        angle_threshold_large = 0.35
        
        for i in range(len(veins)):
            if used[i]:
                continue
            
            current_group = [i]
            used[i] = True
            
            for j in range(i + 1, len(veins)):
                if used[j]:
                    continue
                
                angle_diff = abs(vein_angles[i] - vein_angles[j])
                if angle_diff > np.pi:
                    angle_diff = 2 * np.pi - angle_diff
                
                radius_ratio = min(vein_radii[i], vein_radii[j]) / max(vein_radii[i], vein_radii[j]) if max(vein_radii[i], vein_radii[j]) > 0 else 0
                
                px1, py1 = veins[i][1]
                px2, py2 = veins[j][1]
                point_dist = np.sqrt((px1 - px2)**2 + (py1 - py2)**2)
                
                should_merge = False
                
                if angle_diff < angle_threshold_small:
                    should_merge = True
                elif angle_diff < angle_threshold_large:
                    if radius_ratio > 0.6 and point_dist < petal_size * 0.15:
                        should_merge = True
                
                if should_merge:
                    current_group.append(j)
                    used[j] = True
            
            if len(current_group) == 1:
                merged.append(veins[current_group[0]])
            else:
                max_idx = max(
                    current_group,
                    key=lambda idx: vein_areas[idx]
                )
                merged.append(veins[max_idx])
        
        return merged

    def _classify_shape(
        self,
        contour: np.ndarray,
        petal_area: float,
        petal_perimeter: float,
        circularity: float,
        serration_count: int,
        vein_count: int,
    ) -> Tuple[str, dict]:
        if petal_area < 500:
            return "未知", {"circular": 0.0, "acute": 0.0, "wavy": 0.0}

        petal_size = np.sqrt(petal_area)
        x, y, w, h = cv2.boundingRect(contour)
        aspect_ratio = max(w, h) / min(w, h) if min(w, h) > 0 else 1.0

        serration_density = 0
        if petal_size > 0:
            serration_density = serration_count / (petal_size * 0.01)

        hull = cv2.convexHull(contour)
        hull_perimeter = cv2.arcLength(hull, True)
        perimeter_ratio = petal_perimeter / hull_perimeter if hull_perimeter > 0 else 1.0

        score_circular = 0.0
        score_acute = 0.0
        score_wavy = 0.0

        score_circular += max(0, (circularity - 0.6)) / 0.4 * 50
        if aspect_ratio < 1.4:
            score_circular += 25
        if serration_density < 1.5:
            score_circular += 25

        score_acute += max(0, (aspect_ratio - 1.4)) / 2.0 * 50
        score_acute += max(0, (0.75 - circularity)) / 0.75 * 30
        if serration_density < 2.0:
            score_acute += 20

        score_wavy += max(0, (perimeter_ratio - 1.05)) / 0.3 * 50
        score_wavy += max(0, serration_density - 1.0) / 3.0 * 30
        if serration_count > 5:
            score_wavy += 20

        scores = {
            "圆形": score_circular,
            "尖形": score_acute,
            "波浪形": score_wavy,
        }

        total = sum(scores.values())
        if total > 0:
            scores = {k: v / total * 100 for k, v in scores.items()}
        else:
            scores = {k: 33.33 for k in scores}

        shape_type = max(scores, key=scores.get)
        return shape_type, scores

    def _calculate_serration(
        self, contour: np.ndarray, petal_area: float
    ) -> Tuple[float, int, List[Tuple[int, int]]]:
        if len(contour) < 10 or petal_area < 100:
            return 0.0, 0, []

        smoothed = self._smooth_contour(contour, kernel_size=1)

        hull = cv2.convexHull(smoothed)
        hull_area = cv2.contourArea(hull)
        serration_area = max(0, hull_area - petal_area)

        serration_points = []
        serration_count = 0

        try:
            hull_indices = cv2.convexHull(smoothed, returnPoints=False)
            
            if hull_indices is not None and len(hull_indices) > 3:
                defects = cv2.convexityDefects(smoothed, hull_indices)
                
                if defects is not None:
                    petal_size = np.sqrt(petal_area)
                    min_depth = petal_size * self.min_serration_depth_ratio
                    min_width = petal_size * self.min_serration_width_ratio
                    
                    for i in range(defects.shape[0]):
                        s, e, f, d = defects[i, 0]
                        depth = d / 256.0
                        
                        if depth < min_depth:
                            continue
                        
                        start = tuple(smoothed[s][0])
                        end = tuple(smoothed[e][0])
                        far = tuple(smoothed[f][0])
                        
                        width = np.sqrt(
                            (end[0] - start[0]) ** 2 + (end[1] - start[1]) ** 2
                        )
                        if width < min_width:
                            continue
                        
                        serration_count += 1
                        serration_points.append(far)
        except Exception:
            pass

        if serration_count == 0:
            epsilon = 0.003 * cv2.arcLength(smoothed, True)
            approx = cv2.approxPolyDP(smoothed, epsilon, True)
            
            if len(approx) > 4:
                approx_points = approx.reshape(-1, 2)
                petal_size = np.sqrt(petal_area)
                min_depth = petal_size * self.min_serration_depth_ratio
                
                for i in range(len(approx_points)):
                    p1 = approx_points[i]
                    p2 = approx_points[(i + 1) % len(approx_points)]
                    p3 = approx_points[(i + 2) % len(approx_points)]

                    angle = self._calculate_angle(p1, p2, p3)

                    if angle < 170:
                        mid_point = (p1 + p3) / 2
                        depth = np.linalg.norm(p2 - mid_point)
                        
                        if depth > min_depth * 0.5:
                            dist = cv2.pointPolygonTest(hull, tuple(p2.tolist()), True)
                            if dist < -min_depth * 0.1:
                                serration_count += 1
                                serration_points.append(tuple(p2.tolist()))

        max_serrations = 200
        if len(serration_points) > max_serrations:
            serration_points = serration_points[:max_serrations]
            serration_count = max_serrations

        return serration_area, serration_count, serration_points

    def _calculate_angle(
        self, p1: np.ndarray, p2: np.ndarray, p3: np.ndarray
    ) -> float:
        v1 = p1 - p2
        v2 = p3 - p2

        cosine = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-10)
        cosine = np.clip(cosine, -1, 1)
        angle = np.arccos(cosine) * 180 / np.pi

        return angle

    def generate_overlay_image(
        self, image_bytes: bytes, result: AnalysisResult
    ) -> bytes:
        img_array = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

        overlay = img.copy()

        if result.contours and len(result.contours) > 0:
            contour = np.array(result.contours[0], dtype=np.int32).reshape(-1, 1, 2)
            cv2.drawContours(overlay, [contour], -1, (0, 255, 0), 2)

        for point in result.vein_points:
            cv2.circle(overlay, (point[0], point[1]), 4, (255, 0, 0), -1)

        for point in result.serration_points:
            cv2.circle(overlay, (point[0], point[1]), 5, (0, 0, 255), -1)

        alpha = 0.7
        cv2.addWeighted(overlay, alpha, img, 1 - alpha, 0, img)

        _, buffer = cv2.imencode(".png", img)
        return buffer.tobytes()


def image_to_base64(image_bytes: bytes) -> str:
    return base64.b64encode(image_bytes).decode("utf-8")


def export_results_to_excel(
    results: List[AnalysisResult],
    filenames: Optional[List[str]] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "花瓣分析结果"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    shape_color_map = {
        "圆形": "93C5FD",
        "尖形": "FCD34D",
        "波浪形": "F87171",
        "未知": "D1D5DB",
    }

    headers = [
        "序号",
        "文件名",
        "花瓣形态",
        "圆形概率(%)",
        "尖形概率(%)",
        "波浪形概率(%)",
        "花瓣面积(px²)",
        "花瓣周长(px)",
        "圆形度",
        "脉络数量(条)",
        "锯齿数量(个)",
        "锯齿总面积(px²)",
        "锯齿占比(%)",
        "图像宽度(px)",
        "图像高度(px)",
        "分析时间",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    if filenames is None:
        filenames = [f"标本_{i+1}.png" for i in range(len(results))]

    shape_stats = {"圆形": 0, "尖形": 0, "波浪形": 0, "未知": 0}

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row_idx, result in enumerate(results):
        row = row_idx + 2
        filename = filenames[row_idx] if row_idx < len(filenames) else f"标本_{row_idx+1}.png"

        serration_ratio = 0.0
        if result.petal_area > 0:
            serration_ratio = (result.serration_area / result.petal_area) * 100

        shape_type = result.shape_type or "未知"
        if shape_type in shape_stats:
            shape_stats[shape_type] += 1
        else:
            shape_stats["未知"] += 1

        shape_fill_color = shape_color_map.get(shape_type, "D1D5DB")
        shape_fill = PatternFill(
            start_color=shape_fill_color, end_color=shape_fill_color, fill_type="solid"
        )

        row_data = [
            row_idx + 1,
            filename,
            shape_type,
            round(result.shape_score.get("圆形", 0), 2),
            round(result.shape_score.get("尖形", 0), 2),
            round(result.shape_score.get("波浪形", 0), 2),
            round(result.petal_area, 2),
            round(result.petal_perimeter, 2),
            round(result.circularity, 4),
            result.vein_count,
            result.serration_count,
            round(result.serration_area, 2),
            round(serration_ratio, 2),
            result.original_width,
            result.original_height,
            now_str,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col_num == 3:
                cell.fill = shape_fill

    column_widths = [8, 22, 12, 14, 14, 14, 16, 16, 12, 14, 14, 16, 14, 14, 14, 20]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    if len(results) > 0:
        ws_stats = wb.create_sheet(title="形态统计")

        stats_headers = ["花瓣形态", "数量", "占比(%)"]
        for col_num, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        total = len(results)
        for row_idx, (shape, count) in enumerate(shape_stats.items()):
            row = row_idx + 2
            percentage = (count / total * 100) if total > 0 else 0
            shape_fill_color = shape_color_map.get(shape, "D1D5DB")
            shape_fill = PatternFill(
                start_color=shape_fill_color, end_color=shape_fill_color, fill_type="solid"
            )

            data_row = [shape, count, round(percentage, 2)]
            for col_num, value in enumerate(data_row, 1):
                cell = ws_stats.cell(row=row, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if col_num == 1:
                    cell.fill = shape_fill

        ws_stats.column_dimensions["A"].width = 14
        ws_stats.column_dimensions["B"].width = 10
        ws_stats.column_dimensions["C"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
